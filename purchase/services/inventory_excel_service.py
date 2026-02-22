import io
from decimal import Decimal, InvalidOperation
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from purchase.models.inventory_item import InventoryItem
from purchase.models.expense_category import ExpenseCategory


class InventoryExcelService:

    HEADERS = [
        "Name",
        "Unit",
        "Current Stock",
        "Low Stock Threshold",
        "Cost Per Unit",
        "Category",
        "Stock Value",
    ]

    VALID_UNITS = InventoryItem.Unit.values()

    @classmethod
    def export_to_excel(cls, restaurant):
        """Export all inventory items to an Excel workbook. Returns bytes."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventory"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            bottom=Side(style="thin", color="E5E7EB"),
        )

        # Write headers
        for col_idx, header in enumerate(cls.HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data
        items = InventoryItem.get_items_for_restaurant(restaurant).select_related("category")
        for row_idx, item in enumerate(items, 2):
            ws.cell(row=row_idx, column=1, value=item.name)
            ws.cell(row=row_idx, column=2, value=item.unit)
            ws.cell(row=row_idx, column=3, value=float(item.current_stock))
            ws.cell(row=row_idx, column=4, value=float(item.low_stock_threshold))
            ws.cell(row=row_idx, column=5, value=float(item.cost_per_unit))
            ws.cell(row=row_idx, column=6, value=item.category.name if item.category else "")
            ws.cell(row=row_idx, column=7, value=float(item.stock_value))

            for col in range(1, len(cls.HEADERS) + 1):
                ws.cell(row=row_idx, column=col).border = thin_border

        # Auto-width columns
        for col_idx, header in enumerate(cls.HEADERS, 1):
            max_length = len(header)
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_length + 4

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    @classmethod
    def import_from_excel(cls, file_bytes, restaurant, user):
        """
        Import inventory items from an uploaded Excel file.
        Expected columns: Name, Unit, Current Stock, Low Stock Threshold, Cost Per Unit, Category
        Returns dict with created, updated, errors counts and error details.
        """
        wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            return {"created": 0, "updated": 0, "skipped": 0, "errors": ["Empty file"]}

        # Find header row
        header_row = [str(h).strip().lower() if h else "" for h in rows[0]]
        required = {"name"}
        if not required.issubset(set(header_row)):
            return {"created": 0, "updated": 0, "skipped": 0, "errors": ["Missing required column: Name"]}

        col_map = {}
        for idx, h in enumerate(header_row):
            if h == "name":
                col_map["name"] = idx
            elif h == "unit":
                col_map["unit"] = idx
            elif h in ("current stock", "current_stock", "stock"):
                col_map["current_stock"] = idx
            elif h in ("low stock threshold", "low_stock_threshold", "threshold"):
                col_map["low_stock_threshold"] = idx
            elif h in ("cost per unit", "cost_per_unit", "cost", "unit cost"):
                col_map["cost_per_unit"] = idx
            elif h == "category":
                col_map["category"] = idx

        created = 0
        updated = 0
        skipped = 0
        errors = []

        for row_num, row in enumerate(rows[1:], 2):
            try:
                name = str(row[col_map["name"]]).strip() if row[col_map["name"]] else ""
                if not name:
                    skipped += 1
                    continue

                unit = "KG"
                if "unit" in col_map and row[col_map["unit"]]:
                    raw_unit = str(row[col_map["unit"]]).strip().upper()
                    if raw_unit in cls.VALID_UNITS:
                        unit = raw_unit
                    else:
                        errors.append(f"Row {row_num}: Invalid unit '{raw_unit}', defaulting to KG")

                current_stock = Decimal("0")
                if "current_stock" in col_map and row[col_map["current_stock"]] is not None:
                    try:
                        current_stock = Decimal(str(row[col_map["current_stock"]]))
                    except InvalidOperation:
                        errors.append(f"Row {row_num}: Invalid stock value, defaulting to 0")

                low_stock_threshold = Decimal("0")
                if "low_stock_threshold" in col_map and row[col_map["low_stock_threshold"]] is not None:
                    try:
                        low_stock_threshold = Decimal(str(row[col_map["low_stock_threshold"]]))
                    except InvalidOperation:
                        pass

                cost_per_unit = Decimal("0")
                if "cost_per_unit" in col_map and row[col_map["cost_per_unit"]] is not None:
                    try:
                        cost_per_unit = Decimal(str(row[col_map["cost_per_unit"]]))
                    except InvalidOperation:
                        pass

                category = None
                if "category" in col_map and row[col_map["category"]]:
                    cat_name = str(row[col_map["category"]]).strip()
                    if cat_name:
                        category = ExpenseCategory.get_or_create_category(cat_name, restaurant)

                item, was_created = InventoryItem.objects.get_or_create(
                    restaurant=restaurant,
                    name=name,
                    defaults={
                        "unit": unit,
                        "current_stock": current_stock,
                        "low_stock_threshold": low_stock_threshold,
                        "cost_per_unit": cost_per_unit,
                        "category": category,
                        "updated_by": user,
                        "is_deleted": False,
                    },
                )
                if was_created:
                    created += 1
                else:
                    if item.is_deleted:
                        item.un_delete()
                    item.unit = unit
                    item.low_stock_threshold = low_stock_threshold
                    item.cost_per_unit = cost_per_unit
                    if category:
                        item.category = category
                    item.updated_by = user
                    item.save()
                    updated += 1

            except Exception as e:
                errors.append(f"Row {row_num}: {str(e)}")
                skipped += 1

        return {
            "created": created,
            "updated": updated,
            "skipped": skipped,
            "errors": errors,
        }
