import uuid
from django.shortcuts import redirect
from django.core.cache import cache
from django.db.models import Count, Q
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from common.decorators import manager_or_above_required, waiter_or_above_required
from common.pagination import paginate_queryset
from django.utils.decorators import method_decorator
from sale.models import Dish
from sale.models.order import Order
from sale.models.course import Course
from sale.serializers import DishSerializer
from django.http import HttpResponse

from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import Workbook, load_workbook


def _invalidate_dish_cache(restaurant):
    """Clear cached dish list for a restaurant."""
    cache.delete(f"dishes_{restaurant.id}_true")
    cache.delete(f"dishes_{restaurant.id}_false")


def get_or_create_course_by_name(course_name, restaurant):
    return Course.get_or_create_course_by_name(course_name, restaurant)


@method_decorator(waiter_or_above_required, name="get")
@method_decorator(manager_or_above_required, name="post")
class DishListCreateAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        prices = request.GET.get("prices", "true").lower() == "true"
        search = request.query_params.get("search", "").strip()

        # Skip cache when searching or paginating
        if not search and not request.query_params.get("page"):
            cache_key = f"dishes_{request.restaurant.id}_{prices}"
            cached = cache.get(cache_key)
            if cached is not None:
                return Response(cached)

        dishes = Dish.get_dishes_for_restaurant(request.restaurant)

        if search:
            dishes = dishes.filter(
                Q(name__icontains=search)
                | Q(course__name__icontains=search)
            )

        if not prices:
            dishes = dishes.values("id", "name", "course__name")
            data = self._format_dish_data(dishes)
        else:
            pagination, data = paginate_queryset(dishes, request, DishSerializer)
            if pagination:
                return Response({"results": data, "pagination": pagination})

        # Cache only full, non-search results
        if not search and not request.query_params.get("page"):
            cache_key = f"dishes_{request.restaurant.id}_{prices}"
            cache.set(cache_key, data, timeout=600)
        return Response(data)

    def post(self, request):
        course_name = request.data.get("course_name")
        if not course_name:
            return Response(
                {"course": "This field is required."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        course = Course.get_or_create_course_by_name(course_name, request.restaurant)
        serializer = DishSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save(
                restaurant=request.restaurant, updated_by=request.user, course=course
            )
            _invalidate_dish_cache(request.restaurant)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def _format_dish_data(self, dishes):
        data = []
        for d in dishes:
            data.append(
                {
                    "id": str(d["id"]),
                    "name": d["name"],
                    "course": ({"name": d.get("course__name")}),
                }
            )
        return data


@method_decorator(manager_or_above_required, name="get")
class DishExportAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        dishes = Dish.get_dishes_for_restaurant(request.restaurant).select_related(
            "course"
        )
        wb = Workbook()
        ws = wb.active
        ws.title = "Dishes"
        ws.append(
            [
                "Dish Name",
                "Course",
                "Restaurant Half Price",
                "Restaurant Full Price",
                "Swiggy Half Price",
                "Swiggy Full Price",
                "Zomato Half Price",
                "Zomato Full Price",
            ]
        )
        for dish in dishes:
            ws.append(
                [
                    dish.name,
                    dish.course.name if dish.course else "",
                    dish.restaurant_half_price,
                    dish.restaurant_full_price,
                    dish.swiggy_half_price,
                    dish.swiggy_full_price,
                    dish.zomato_half_price,
                    dish.zomato_full_price,
                ]
            )
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="dishes_menu.xlsx"'
        wb.save(response)
        return response


@method_decorator(waiter_or_above_required, name="get")
@method_decorator(manager_or_above_required, name="put")
@method_decorator(manager_or_above_required, name="delete")
class DishDetailUpdateDeleteAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get_object(self, id: uuid.UUID):
        return Dish.get_by_id(id)

    def get(self, request, id):
        dish = self.get_object(id)
        serializer = DishSerializer(dish)
        return Response(serializer.data)

    def put(self, request, id):
        dish = self.get_object(id)
        if not dish:
            return Response(
                {"detail": "Dish not found."}, status=status.HTTP_404_NOT_FOUND
            )
        course_name = request.data.get("course_name")
        if course_name:
            course = Course.get_or_create_course_by_name(
                course_name, request.restaurant
            )
        else:
            course = dish.course
        serializer = DishSerializer(dish, data=request.data)
        if serializer.is_valid():
            serializer.save(updated_by=request.user, course=course)
            _invalidate_dish_cache(request.restaurant)
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def delete(self, request, id):
        dish = self.get_object(id)
        if not dish:
            return Response(
                {"detail": "Dish not found."}, status=status.HTTP_404_NOT_FOUND
            )
        dish.soft_delete(request.user)
        _invalidate_dish_cache(request.restaurant)
        return Response(status=status.HTTP_204_NO_CONTENT)


@method_decorator(manager_or_above_required, name="post")
class DishImportAPIView(APIView):
    permission_classes = [IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"error": "No file uploaded."}, status=400)
        wb = load_workbook(file)
        ws = wb.active
        headers = [
            str(cell.value).strip() if cell.value is not None else ""
            for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_dict = dict(zip(headers, row))
            name_val = row_dict.get("Dish Name")
            course_val = row_dict.get("Course")
            name = str(name_val).strip() if name_val is not None else ""
            course_name = str(course_val).strip() if course_val is not None else ""
            if not name:
                continue
            course = (
                get_or_create_course_by_name(course_name, request.restaurant)
                if course_name
                else None
            )
            dish_obj, created = Dish.objects.update_or_create(
                restaurant=request.restaurant,
                name=name,
                is_deleted=False,
                defaults={
                    "restaurant_half_price": row_dict.get("Restaurant Half Price") or 0,
                    "restaurant_full_price": row_dict.get("Restaurant Full Price") or 0,
                    "swiggy_half_price": row_dict.get("Swiggy Half Price") or 0,
                    "swiggy_full_price": row_dict.get("Swiggy Full Price") or 0,
                    "zomato_half_price": row_dict.get("Zomato Half Price") or 0,
                    "zomato_full_price": row_dict.get("Zomato Full Price") or 0,
                    "updated_by": request.user,
                },
            )
            # Ensure course is updated if dish already exists and course is different
            if dish_obj.course != course:
                dish_obj.course = course
                dish_obj.save(update_fields=["course"])
        _invalidate_dish_cache(request.restaurant)
        return redirect("menu")


@method_decorator(waiter_or_above_required, name="get")
class PopularDishesAPIView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        popular = (
            Order.objects.filter(
                restaurant=request.restaurant,
                is_deleted=False,
                dish__is_deleted=False,
                dish__isnull=False,
            )
            .values("dish_id")
            .annotate(order_count=Count("id"))
            .order_by("-order_count")[:10]
        )
        return Response(list(popular))
