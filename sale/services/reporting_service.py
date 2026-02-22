from datetime import datetime
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sale.models.bill import Bill
from sale.models.order import Order


class ReportingService:
    @classmethod
    def get_filtered_invoices(cls, start_date, end_date, order_type, payment_type, restaurant):
        """Filter invoices based on date range and order/payment types"""
        bills = Bill.objects.filter(restaurant=restaurant, invoice_number__isnull=False)

        payment_type = None if payment_type == "ALL" else payment_type
        order_type = None if order_type == "ALL" else order_type

        # Build the filter dictionary
        filters = {}

        # Apply date filters - now handled by _get_date_filters
        date_filters = cls._get_date_filters(start_date, end_date)
        if date_filters is None:
            return Bill.objects.none()

        filters.update(date_filters)

        # Add additional filters if provided
        if order_type:
            filters["order_type"] = order_type
        if payment_type:
            filters["payment_type"] = payment_type

        # Apply all filters at once
        return bills.filter(**filters)

    @staticmethod
    def generate_excel_response(bills):
        """Generate GSTR-1 compliant Excel response using openpyxl library.
        Returns a tuple of (excel_content, filename)
        """
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "GSTR-1 Report"

        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # GSTR-1 compliant headers
        headers = [
            "Invoice Date",
            "Invoice Number",
            "Customer Name",
            "Customer GSTIN",
            "Place of Supply",
            "HSN/SAC Code",
            "Taxable Value",
            "CGST Rate (%)",
            "CGST Amount",
            "SGST Rate (%)",
            "SGST Amount",
            "IGST Rate (%)",
            "IGST Amount",
            "Total Tax Amount",
            "Total Invoice Value",
            "Invoice Type",
            "Is Deleted",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        if bills:
            for row, bill in enumerate(bills, 2):
                # Calculate tax rates (assuming 2.5% for CGST/SGST, 5% for IGST)
                cgst_rate = 2.5 if bill.cgst and bill.cgst > 0 else 0
                sgst_rate = 2.5 if bill.sgst and bill.sgst > 0 else 0
                igst_rate = 5.0 if bill.igst and bill.igst > 0 else 0

                # Calculate total tax amount
                total_tax = (bill.cgst or 0) + (bill.sgst or 0) + (bill.igst or 0)

                # Determine invoice type based on customer GSTIN
                invoice_type = "B2B" if bill.customer_gstin else "B2C"

                ws.cell(row=row, column=1, value=bill.date.strftime("%d/%m/%Y") if bill.date else "")
                ws.cell(row=row, column=2, value=bill.full_invoice_number)
                ws.cell(row=row, column=3, value=bill.customer_name if bill.customer_name else "")
                ws.cell(row=row, column=4, value=bill.customer_gstin if bill.customer_gstin else "")
                ws.cell(row=row, column=5, value=bill.customer_gstin[:2] if bill.customer_gstin else "")
                ws.cell(row=row, column=6, value="996331")
                ws.cell(row=row, column=7, value=float(bill.net) if bill.net is not None else 0.00)
                ws.cell(row=row, column=8, value=cgst_rate)
                ws.cell(row=row, column=9, value=float(bill.cgst) if bill.cgst is not None else 0.00)
                ws.cell(row=row, column=10, value=sgst_rate)
                ws.cell(row=row, column=11, value=float(bill.sgst) if bill.sgst is not None else 0.00)
                ws.cell(row=row, column=12, value=igst_rate)
                ws.cell(row=row, column=13, value=float(bill.igst) if bill.igst is not None else 0.00)
                ws.cell(row=row, column=14, value=float(total_tax))
                ws.cell(row=row, column=15, value=float(bill.amount) if bill.amount is not None else 0.00)
                ws.cell(row=row, column=16, value=invoice_type)
                ws.cell(row=row, column=17, value="Yes" if bill.is_deleted else "No")

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Add summary row with totals
        if bills:
            summary_row = len(bills) + 3
            ws.cell(row=summary_row, column=1, value="TOTAL")
            ws.cell(row=summary_row, column=1).font = Font(bold=True)

            # Calculate totals
            total_taxable = sum(float(bill.net) if bill.net else 0 for bill in bills)
            total_cgst = sum(float(bill.cgst) if bill.cgst else 0 for bill in bills)
            total_sgst = sum(float(bill.sgst) if bill.sgst else 0 for bill in bills)
            total_igst = sum(float(bill.igst) if bill.igst else 0 for bill in bills)
            total_tax_amount = total_cgst + total_sgst + total_igst
            total_invoice_value = sum(float(bill.amount) if bill.amount else 0 for bill in bills)

            ws.cell(row=summary_row, column=7, value=total_taxable)
            ws.cell(row=summary_row, column=9, value=total_cgst)
            ws.cell(row=summary_row, column=11, value=total_sgst)
            ws.cell(row=summary_row, column=13, value=total_igst)
            ws.cell(row=summary_row, column=14, value=total_tax_amount)
            ws.cell(row=summary_row, column=15, value=total_invoice_value)

            # Style summary row
            for col in range(1, 21):
                cell = ws.cell(row=summary_row, column=col)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"GSTR1_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return output.getvalue(), filename

    @staticmethod
    def generate_dish_summary_report(bills):
        """Generate dish summary report showing dish count and invoice numbers.
        Returns a tuple of (excel_content, filename)
        """
        # Get all orders from these bills
        orders = Order.objects.filter(
            bill__in=bills
        ).select_related('dish', 'bill').values(
            'dish__name',
            'size',
            'quantity',
            'bill__invoice_number'
        )

        # Group dishes and collect invoice numbers
        dish_data = {}
        for order in orders:
            dish_name = order['dish__name']
            size = order['size']
            dish_key = f"{dish_name.lower()}_{size.lower()}"  # Use lowercase for case-insensitive grouping
            quantity = order['quantity']
            invoice_number = order['bill__invoice_number']

            if dish_key not in dish_data:
                dish_data[dish_key] = {
                    'dish_name': dish_name,  # Store original name for display
                    'size': size.upper(),  # Store original size for display
                    'total_count': 0,
                    'invoice_numbers': set()
                }

            dish_data[dish_key]['total_count'] += quantity
            dish_data[dish_key]['invoice_numbers'].add(str(invoice_number))

        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Dish Summary"

        # Define header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Headers
        headers = ["Dish", "Size", "Count", "Invoices"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Write data rows
        row = 2
        for dish_key, data in sorted(dish_data.items(), key=lambda x: x[1]['total_count'], reverse=True):
            # Convert invoice numbers set to comma-separated string
            invoice_numbers_csv = ", ".join(sorted(data['invoice_numbers']))

            ws.cell(row=row, column=1, value=data['dish_name'])  # Use original name for display
            ws.cell(row=row, column=2, value=data['size'])  # Size column
            ws.cell(row=row, column=3, value=data['total_count'])
            ws.cell(row=row, column=4, value=invoice_numbers_csv)
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Increased max width for invoice numbers
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"Dish_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return output.getvalue(), filename

    @staticmethod
    def _get_date_filters(start_date, end_date):
        """Parse and validate date filters"""
        filters = {}

        # If no dates are provided, use today's date
        if not start_date and not end_date:
            filters["created_at__date"] = datetime.now().date()
            return filters

        if start_date:
            try:
                datetime.strptime(start_date, "%Y-%m-%d")
                filters["created_at__date__gte"] = start_date
            except (ValueError, TypeError):
                return None

        if end_date:
            try:
                datetime.strptime(end_date, "%Y-%m-%d")
                filters["created_at__date__lte"] = end_date
            except (ValueError, TypeError):
                return None

        return filters
